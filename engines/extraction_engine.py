# -*- coding: utf-8 -*-
"""
محرك استخراج المحتوى - استخراج النص والبيانات من الملفات
"""

import logging
from typing import Dict, Optional

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from pptx import Presentation
except ImportError:
    Presentation = None

try:
    from PIL import Image
except ImportError:
    Image = None

logger = logging.getLogger(__name__)


class ExtractionEngine:
    """محرك استخراج المحتوى من الملفات"""

    def __init__(self):
        """تهيئة محرك الاستخراج"""
        pass

    def extract_from_pdf(self, file_path: str) -> Dict:
        """
        استخراج النص من ملف PDF
        
        Args:
            file_path: مسار ملف PDF
            
        Returns:
            قاموس بالمحتوى والبيانات الوصفية
        """
        result = {
            'success': False,
            'text': '',
            'metadata': {},
            'page_count': 0,
            'error': None
        }

        try:
            if not fitz:
                result['error'] = "مكتبة PyMuPDF غير مثبتة"
                return result

            doc = fitz.open(file_path)
            result['page_count'] = doc.page_count

            # استخراج النص من جميع الصفحات
            full_text = ""
            for page_num in range(doc.page_count):
                page = doc[page_num]
                text = page.get_text()
                full_text += text + "\n"

            result['text'] = full_text.strip()

            # استخراج البيانات الوصفية
            if doc.metadata:
                result['metadata'] = {
                    'title': doc.metadata.get('title', ''),
                    'author': doc.metadata.get('author', ''),
                    'subject': doc.metadata.get('subject', ''),
                    'creator': doc.metadata.get('creator', ''),
                    'producer': doc.metadata.get('producer', ''),
                    'creation_date': str(doc.metadata.get('creationDate', '')),
                    'modification_date': str(doc.metadata.get('modDate', '')),
                }

            result['success'] = True
            logger.info(f"تم استخراج النص من PDF بنجاح ({result['page_count']} صفحات)")

        except Exception as e:
            result['error'] = f"خطأ في استخراج PDF: {str(e)}"
            logger.error(result['error'])

        return result

    def extract_from_docx(self, file_path: str) -> Dict:
        """
        استخراج النص من ملف DOCX
        
        Args:
            file_path: مسار ملف DOCX
            
        Returns:
            قاموس بالمحتوى والبيانات الوصفية
        """
        result = {
            'success': False,
            'text': '',
            'metadata': {},
            'paragraph_count': 0,
            'error': None
        }

        try:
            if not Document:
                result['error'] = "مكتبة python-docx غير مثبتة"
                return result

            doc = Document(file_path)
            
            # استخراج جميع الفقرات
            paragraphs = []
            for para in doc.paragraphs:
                if para.text.strip():
                    paragraphs.append(para.text)
            
            result['text'] = '\n'.join(paragraphs)
            result['paragraph_count'] = len(paragraphs)
            
            # استخراج البيانات الوصفية
            if doc.core_properties:
                props = doc.core_properties
                result['metadata'] = {
                    'title': props.title or '',
                    'author': props.author or '',
                    'subject': props.subject or '',
                    'keywords': props.keywords or '',
                    'created': str(props.created) if props.created else '',
                    'modified': str(props.modified) if props.modified else '',
                }

            result['success'] = True
            logger.info(f"تم استخراج النص من DOCX بنجاح ({result['paragraph_count']} فقرات)")

        except Exception as e:
            result['error'] = f"خطأ في استخراج DOCX: {str(e)}"
            logger.error(result['error'])

        return result

    def extract_from_xlsx(self, file_path: str) -> Dict:
        """
        استخراج البيانات من ملف XLSX
        
        Args:
            file_path: مسار ملف XLSX
            
        Returns:
            قاموس بالمحتوى والبيانات الوصفية
        """
        result = {
            'success': False,
            'text': '',
            'metadata': {},
            'sheet_count': 0,
            'error': None
        }

        try:
            if not openpyxl:
                result['error'] = "مكتبة openpyxl غير مثبتة"
                return result

            wb = openpyxl.load_workbook(file_path)
            result['sheet_count'] = len(wb.sheetnames)
            
            # استخراج البيانات من جميع الأوراق
            all_text = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_text = [f"[ورقة: {sheet_name}]"]
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' | '.join([str(cell) if cell is not None else '' for cell in row])
                    if row_text.strip():
                        sheet_text.append(row_text)
                all_text.extend(sheet_text)
            
            result['text'] = '\n'.join(all_text)
            result['success'] = True
            logger.info(f"تم استخراج البيانات من XLSX بنجاح ({result['sheet_count']} أوراق)")

        except Exception as e:
            result['error'] = f"خطأ في استخراج XLSX: {str(e)}"
            logger.error(result['error'])

        return result

    def extract_from_image(self, file_path: str) -> Dict:
        """
        استخراج معلومات من ملف صورة
        
        Args:
            file_path: مسار ملف الصورة
            
        Returns:
            قاموس بمعلومات الصورة
        """
        result = {
            'success': False,
            'text': '',
            'metadata': {},
            'error': None,
            'requires_ocr': True
        }

        try:
            if not Image:
                result['error'] = "مكتبة Pillow غير مثبتة"
                return result

            img = Image.open(file_path)
            result['metadata'] = {
                'format': img.format,
                'width': img.width,
                'height': img.height,
                'mode': img.mode,
                'size': f"{img.width}x{img.height}",
            }
            
            # نص افتراضي (سيتم استخراجه عبر OCR)
            result['text'] = "[صورة - يتطلب معالجة OCR]"
            result['success'] = True
            logger.info(f"تم استخراج معلومات الصورة: {result['metadata']['size']}")

        except Exception as e:
            result['error'] = f"خطأ في استخراج معلومات الصورة: {str(e)}"
            logger.error(result['error'])

        return result

    def extract_content(self, file_path: str, file_type: str) -> Dict:
        """
        استخراج المحتوى من الملف حسب نوعه
        
        Args:
            file_path: مسار الملف
            file_type: نوع الملف
            
        Returns:
            قاموس بالمحتوى المستخرج
        """
        result = {
            'success': False,
            'text': '',
            'metadata': {},
            'requires_ocr': False,
            'error': None
        }

        try:
            if file_type.lower() == 'pdf':
                result = self.extract_from_pdf(file_path)
            elif file_type.lower() == 'word':
                result = self.extract_from_docx(file_path)
            elif file_type.lower() == 'excel':
                result = self.extract_from_xlsx(file_path)
            elif file_type.lower() == 'image':
                result = self.extract_from_image(file_path)
            else:
                result['error'] = f"نوع الملف غير مدعوم: {file_type}"
                logger.warning(result['error'])

            # تحديد ما إذا كان النص فارغاً ويتطلب OCR
            if result.get('success') and not result.get('text', '').strip():
                result['requires_ocr'] = True
                logger.info("النص فارغ - سيتم تطبيق OCR")

        except Exception as e:
            result['error'] = f"خطأ في استخراج المحتوى: {str(e)}"
            logger.error(result['error'])

        return result
